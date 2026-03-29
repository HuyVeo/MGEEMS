
import sys
import os
from collections import Counter, defaultdict

try:
    from openpyxl import load_workbook
except ImportError:
    print("Cần cài: pip install openpyxl")
    sys.exit(1)

try:
    import matplotlib
    matplotlib.use('Agg')
    import matplotlib.pyplot as plt
    import matplotlib.patches as mpatches
    import numpy as np
except ImportError:
    print("Cần cài: pip install matplotlib numpy")
    sys.exit(1)

# ============================================================
# CẤU HÌNH CHUNG
# ============================================================
plt.rcParams.update({
    'font.size': 12,
    'axes.titlesize': 14,
    'axes.labelsize': 12,
    'figure.dpi': 150,
    'savefig.bbox': 'tight',
    'savefig.pad_inches': 0.3,
})

SCORE_COLORS = {
    0: '#E74C3C',   # đỏ
    1: '#E67E22',   # cam
    2: '#F1C40F',   # vàng
    3: '#3498DB',   # xanh dương
    4: '#2ECC71',   # xanh lá
}
SCORE_LABELS = {
    0: 'Unrelated',
    1: 'Slightly Related',
    2: 'Related',
    3: 'Similar',
    4: 'Very Similar',
}

DOMAIN_COLORS = {
    'Y tế':       '#E74C3C',
    'Đời sống':   '#3498DB',
    'Sinh học':    '#2ECC71',
    'Chính trị':  '#9B59B6',
    'Kinh tế':    '#F39C12',
    'Nông nghiệp':'#1ABC9C',
    'Ngôn ngữ':   '#E67E22',
    'Kỹ thuật':   '#34495E',
    'CNTT':        '#2980B9',
}


def read_dataset(filepath):
    """Đọc dataset từ file xlsx."""
    wb = load_workbook(filepath)
    ws = wb.active
    
    data = []
    for row in range(2, ws.max_row + 1):
        entry = {
            'concept1': ws.cell(row=row, column=2).value or '',
            'concept2': ws.cell(row=row, column=3).value or '',
            'domain1':  ws.cell(row=row, column=4).value or '',
            'domain2':  ws.cell(row=row, column=5).value or '',
            'score':    ws.cell(row=row, column=6).value,
        }
        if entry['score'] is not None:
            entry['score'] = int(entry['score'])
            data.append(entry)
    
    return data


def plot_score_distribution(data, output_dir):
    """Biểu đồ 1: Phân bố Score (Bar chart + Pie chart)."""
    fig, (ax1, ax2) = plt.subplots(1, 2, figsize=(14, 5))
    
    sc = Counter(p['score'] for p in data)
    total = len(data)
    scores = sorted(sc.keys())
    counts = [sc[s] for s in scores]
    colors = [SCORE_COLORS[s] for s in scores]
    labels = [f"Score {s}\n{SCORE_LABELS[s]}" for s in scores]
    
    # Bar chart
    bars = ax1.bar(range(len(scores)), counts, color=colors, edgecolor='white', linewidth=1.5)
    ax1.set_xticks(range(len(scores)))
    ax1.set_xticklabels([f"Score {s}" for s in scores])
    ax1.set_ylabel('Số cặp câu')
    ax1.set_title('Phân bố Similarity Score')
    ax1.spines['top'].set_visible(False)
    ax1.spines['right'].set_visible(False)
    
    for bar, count in zip(bars, counts):
        pct = count / total * 100
        ax1.text(bar.get_x() + bar.get_width()/2, bar.get_height() + 1,
                f'{count}\n({pct:.1f}%)', ha='center', va='bottom', fontsize=10, fontweight='bold')
    
    ax1.set_ylim(0, max(counts) * 1.25)
    
    # Pie chart
    wedges, texts, autotexts = ax2.pie(
        counts, labels=None, colors=colors, autopct='%1.1f%%',
        startangle=90, pctdistance=0.75, wedgeprops=dict(width=0.5, edgecolor='white')
    )
    for t in autotexts:
        t.set_fontsize(10)
        t.set_fontweight('bold')
    
    ax2.set_title('Tỷ lệ Score')
    legend_labels = [f"Score {s} - {SCORE_LABELS[s]} ({sc[s]})" for s in scores]
    ax2.legend(wedges, legend_labels, loc='center left', bbox_to_anchor=(1, 0.5), fontsize=9)
    
    plt.suptitle(f'ViConSim Dataset - Phân bố Score (n={total})', fontsize=16, fontweight='bold', y=1.02)
    plt.tight_layout()
    plt.savefig(os.path.join(output_dir, '1_score_distribution.png'))
    plt.close()
    print("  ✅ 1_score_distribution.png")


def plot_domain_distribution(data, output_dir):
    """Biểu đồ 2: Phân bố theo lĩnh vực (Horizontal bar chart)."""
    fig, ax = plt.subplots(figsize=(12, 6))
    
    domain_count = Counter()
    for p in data:
        domain_count[p['domain1']] += 1
        domain_count[p['domain2']] += 1
    
    domains = sorted(domain_count.keys(), key=lambda x: domain_count[x])
    counts = [domain_count[d] for d in domains]
    colors = [DOMAIN_COLORS.get(d, '#95A5A6') for d in domains]
    total = sum(counts)
    
    bars = ax.barh(range(len(domains)), counts, color=colors, edgecolor='white', linewidth=1.5, height=0.6)
    ax.set_yticks(range(len(domains)))
    ax.set_yticklabels(domains, fontsize=11)
    ax.set_xlabel('Số lần xuất hiện (Domain 1 + Domain 2)')
    ax.set_title(f'Phân bố theo Lĩnh vực (n={total} xuất hiện, {len(domains)} lĩnh vực)', fontsize=14, fontweight='bold')
    ax.spines['top'].set_visible(False)
    ax.spines['right'].set_visible(False)
    
    for bar, count in zip(bars, counts):
        pct = count / total * 100
        ax.text(bar.get_width() + 2, bar.get_y() + bar.get_height()/2,
                f'{count} ({pct:.1f}%)', ha='left', va='center', fontsize=10, fontweight='bold')
    
    ax.set_xlim(0, max(counts) * 1.2)
    plt.tight_layout()
    plt.savefig(os.path.join(output_dir, '2_domain_distribution.png'))
    plt.close()
    print("  ✅ 2_domain_distribution.png")


def plot_domain_pie(data, output_dir):
    """Biểu đồ 3: Pie chart lĩnh vực."""
    fig, ax = plt.subplots(figsize=(9, 7))
    
    domain_count = Counter()
    for p in data:
        domain_count[p['domain1']] += 1
        domain_count[p['domain2']] += 1
    
    domains = sorted(domain_count.keys(), key=lambda x: -domain_count[x])
    counts = [domain_count[d] for d in domains]
    colors = [DOMAIN_COLORS.get(d, '#95A5A6') for d in domains]
    
    explode = [0.05] * len(domains)
    explode[0] = 0.1  # tách domain lớn nhất
    
    wedges, texts, autotexts = ax.pie(
        counts, labels=None, colors=colors, autopct='%1.1f%%',
        startangle=140, explode=explode, pctdistance=0.8,
        wedgeprops=dict(edgecolor='white', linewidth=2)
    )
    for t in autotexts:
        t.set_fontsize(9)
        t.set_fontweight('bold')
    
    legend_labels = [f"{d} ({domain_count[d]})" for d in domains]
    ax.legend(wedges, legend_labels, loc='center left', bbox_to_anchor=(1, 0.5), fontsize=10)
    ax.set_title(f'Tỷ lệ Lĩnh vực trong ViConSim ({len(domains)} lĩnh vực)', 
                 fontsize=14, fontweight='bold')
    
    plt.tight_layout()
    plt.savefig(os.path.join(output_dir, '3_domain_pie.png'))
    plt.close()
    print("  ✅ 3_domain_pie.png")


def plot_score_by_domain_type(data, output_dir):
    """Biểu đồ 4: Score phân theo cùng/khác lĩnh vực (Grouped bar)."""
    fig, (ax1, ax2) = plt.subplots(1, 2, figsize=(14, 5))
    
    same = [p['score'] for p in data if p['domain1'] == p['domain2']]
    cross = [p['score'] for p in data if p['domain1'] != p['domain2']]
    
    # Grouped bar chart
    sc_same = Counter(same)
    sc_cross = Counter(cross)
    scores = [0, 1, 2, 3, 4]
    
    x = np.arange(len(scores))
    width = 0.35
    
    bars1 = ax1.bar(x - width/2, [sc_same.get(s, 0) for s in scores], width,
                    label=f'Cùng lĩnh vực (n={len(same)})', color='#3498DB', edgecolor='white')
    bars2 = ax1.bar(x + width/2, [sc_cross.get(s, 0) for s in scores], width,
                    label=f'Khác lĩnh vực (n={len(cross)})', color='#E74C3C', edgecolor='white')
    
    ax1.set_xticks(x)
    ax1.set_xticklabels([f"Score {s}" for s in scores])
    ax1.set_ylabel('Số cặp câu')
    ax1.set_title('Phân bố Score theo loại cặp Domain')
    ax1.legend(fontsize=10)
    ax1.spines['top'].set_visible(False)
    ax1.spines['right'].set_visible(False)
    
    # Thêm số trên cột
    for bars in [bars1, bars2]:
        for bar in bars:
            h = bar.get_height()
            if h > 0:
                ax1.text(bar.get_x() + bar.get_width()/2, h + 0.5,
                        str(int(h)), ha='center', va='bottom', fontsize=9)
    
    # Pie chart: same vs cross
    sizes = [len(same), len(cross)]
    labels_pie = [f'Cùng lĩnh vực\n({len(same)} cặp)', f'Khác lĩnh vực\n({len(cross)} cặp)']
    colors_pie = ['#3498DB', '#E74C3C']
    
    wedges, texts, autotexts = ax2.pie(
        sizes, labels=labels_pie, colors=colors_pie, autopct='%1.1f%%',
        startangle=90, wedgeprops=dict(width=0.5, edgecolor='white'),
        pctdistance=0.75
    )
    for t in autotexts:
        t.set_fontsize(11)
        t.set_fontweight('bold')
    
    ax2.set_title('Tỷ lệ cùng/khác lĩnh vực')
    
    # Thêm score TB
    avg_same = sum(same) / len(same) if same else 0
    avg_cross = sum(cross) / len(cross) if cross else 0
    ax2.text(0, -1.3, f'Score TB cùng LV: {avg_same:.2f} | Score TB khác LV: {avg_cross:.2f}',
             ha='center', fontsize=11, fontstyle='italic',
             bbox=dict(boxstyle='round,pad=0.3', facecolor='lightyellow', alpha=0.8))
    
    plt.suptitle('ViConSim - Phân tích Same vs Cross Domain', fontsize=16, fontweight='bold', y=1.02)
    plt.tight_layout()
    plt.savefig(os.path.join(output_dir, '4_score_by_domain_type.png'))
    plt.close()
    print("  ✅ 4_score_by_domain_type.png")


def plot_heatmap_cross_domain(data, output_dir):
    """Biểu đồ 5: Heatmap số cặp giữa các domain."""
    fig, ax = plt.subplots(figsize=(10, 8))
    
    # Đếm số cặp giữa các domain
    all_domains = sorted(set(p['domain1'] for p in data) | set(p['domain2'] for p in data))
    n = len(all_domains)
    matrix = np.zeros((n, n))
    
    domain_idx = {d: i for i, d in enumerate(all_domains)}
    
    for p in data:
        i = domain_idx[p['domain1']]
        j = domain_idx[p['domain2']]
        matrix[i][j] += 1
        if i != j:
            matrix[j][i] += 1  # symmetric
    
    # Plot heatmap
    im = ax.imshow(matrix, cmap='YlOrRd', interpolation='nearest')
    
    ax.set_xticks(range(n))
    ax.set_yticks(range(n))
    ax.set_xticklabels(all_domains, rotation=45, ha='right', fontsize=10)
    ax.set_yticklabels(all_domains, fontsize=10)
    
    # Thêm số vào ô
    for i in range(n):
        for j in range(n):
            val = int(matrix[i][j])
            if val > 0:
                color = 'white' if val > matrix.max() * 0.6 else 'black'
                ax.text(j, i, str(val), ha='center', va='center',
                       fontsize=10, fontweight='bold', color=color)
    
    ax.set_title('Heatmap: Số cặp câu giữa các lĩnh vực', fontsize=14, fontweight='bold')
    
    cbar = plt.colorbar(im, ax=ax, shrink=0.8)
    cbar.set_label('Số cặp câu', fontsize=11)
    
    plt.tight_layout()
    plt.savefig(os.path.join(output_dir, '5_domain_heatmap.png'))
    plt.close()
    print("  ✅ 5_domain_heatmap.png")


def plot_score_heatmap_by_domain(data, output_dir):
    """Biểu đồ 6: Score trung bình giữa các cặp domain."""
    fig, ax = plt.subplots(figsize=(10, 8))
    
    all_domains = sorted(set(p['domain1'] for p in data) | set(p['domain2'] for p in data))
    n = len(all_domains)
    matrix = np.full((n, n), np.nan)
    count_matrix = np.zeros((n, n))
    sum_matrix = np.zeros((n, n))
    
    domain_idx = {d: i for i, d in enumerate(all_domains)}
    
    for p in data:
        i = domain_idx[p['domain1']]
        j = domain_idx[p['domain2']]
        sum_matrix[i][j] += p['score']
        count_matrix[i][j] += 1
        if i != j:
            sum_matrix[j][i] += p['score']
            count_matrix[j][i] += 1
    
    for i in range(n):
        for j in range(n):
            if count_matrix[i][j] > 0:
                matrix[i][j] = sum_matrix[i][j] / count_matrix[i][j]
    
    # Mask NaN
    masked = np.ma.masked_invalid(matrix)
    
    im = ax.imshow(masked, cmap='RdYlGn', vmin=0, vmax=4, interpolation='nearest')
    
    ax.set_xticks(range(n))
    ax.set_yticks(range(n))
    ax.set_xticklabels(all_domains, rotation=45, ha='right', fontsize=10)
    ax.set_yticklabels(all_domains, fontsize=10)
    
    for i in range(n):
        for j in range(n):
            if not np.isnan(matrix[i][j]):
                val = matrix[i][j]
                color = 'white' if val < 1.5 or val > 3.0 else 'black'
                ax.text(j, i, f'{val:.1f}', ha='center', va='center',
                       fontsize=10, fontweight='bold', color=color)
    
    ax.set_title('Score trung bình giữa các cặp lĩnh vực', fontsize=14, fontweight='bold')
    
    cbar = plt.colorbar(im, ax=ax, shrink=0.8)
    cbar.set_label('Score trung bình (0-4)', fontsize=11)
    cbar.set_ticks([0, 1, 2, 3, 4])
    
    plt.tight_layout()
    plt.savefig(os.path.join(output_dir, '6_avg_score_heatmap.png'))
    plt.close()
    print("  ✅ 6_avg_score_heatmap.png")


def plot_summary_dashboard(data, output_dir):
    """Biểu đồ 7: Dashboard tổng hợp."""
    fig = plt.figure(figsize=(16, 10))
    fig.suptitle('ViConSim Dataset - Dashboard Tổng Hợp', fontsize=18, fontweight='bold', y=0.98)
    
    # --- Panel 1: Score distribution (top-left) ---
    ax1 = fig.add_subplot(2, 3, 1)
    sc = Counter(p['score'] for p in data)
    scores = sorted(sc.keys())
    counts = [sc[s] for s in scores]
    colors = [SCORE_COLORS[s] for s in scores]
    ax1.bar(scores, counts, color=colors, edgecolor='white')
    ax1.set_title('Phân bố Score', fontweight='bold')
    ax1.set_xlabel('Score')
    ax1.set_ylabel('Số cặp')
    for s, c in zip(scores, counts):
        ax1.text(s, c + 1, str(c), ha='center', fontsize=9, fontweight='bold')
    ax1.spines['top'].set_visible(False)
    ax1.spines['right'].set_visible(False)
    
    # --- Panel 2: Domain pie (top-center) ---
    ax2 = fig.add_subplot(2, 3, 2)
    dc = Counter()
    for p in data:
        dc[p['domain1']] += 1
        dc[p['domain2']] += 1
    domains_sorted = sorted(dc.keys(), key=lambda x: -dc[x])
    d_counts = [dc[d] for d in domains_sorted]
    d_colors = [DOMAIN_COLORS.get(d, '#95A5A6') for d in domains_sorted]
    wedges, _, autotexts = ax2.pie(d_counts, colors=d_colors, autopct='%1.0f%%',
                                    startangle=140, pctdistance=0.8,
                                    wedgeprops=dict(edgecolor='white'))
    for t in autotexts:
        t.set_fontsize(8)
    ax2.set_title(f'{len(domains_sorted)} Lĩnh vực', fontweight='bold')
    ax2.legend(wedges, domains_sorted, loc='center left', bbox_to_anchor=(1, 0.5), fontsize=8)
    
    # --- Panel 3: Same vs Cross (top-right) ---
    ax3 = fig.add_subplot(2, 3, 3)
    same = sum(1 for p in data if p['domain1'] == p['domain2'])
    cross = len(data) - same
    ax3.pie([same, cross], labels=[f'Cùng LV\n{same}', f'Khác LV\n{cross}'],
            colors=['#3498DB', '#E74C3C'], autopct='%1.1f%%',
            wedgeprops=dict(width=0.5, edgecolor='white'), startangle=90)
    ax3.set_title('Cùng vs Khác Lĩnh vực', fontweight='bold')
    
    # --- Panel 4: Score boxplot by domain type (bottom-left) ---
    ax4 = fig.add_subplot(2, 3, 4)
    same_scores = [p['score'] for p in data if p['domain1'] == p['domain2']]
    cross_scores = [p['score'] for p in data if p['domain1'] != p['domain2']]
    bp = ax4.boxplot([same_scores, cross_scores], labels=['Cùng LV', 'Khác LV'],
                     patch_artist=True, widths=0.5)
    bp['boxes'][0].set_facecolor('#3498DB')
    bp['boxes'][1].set_facecolor('#E74C3C')
    for b in bp['boxes']:
        b.set_alpha(0.7)
    ax4.set_ylabel('Score')
    ax4.set_title('Score theo loại cặp', fontweight='bold')
    avg_s = sum(same_scores)/len(same_scores) if same_scores else 0
    avg_c = sum(cross_scores)/len(cross_scores) if cross_scores else 0
    ax4.text(1, -0.8, f'TB: {avg_s:.2f}', ha='center', fontsize=9, color='#3498DB', fontweight='bold')
    ax4.text(2, -0.8, f'TB: {avg_c:.2f}', ha='center', fontsize=9, color='#E74C3C', fontweight='bold')
    ax4.spines['top'].set_visible(False)
    ax4.spines['right'].set_visible(False)
    
    # --- Panel 5: Stacked bar - Score distribution per domain (bottom-center+right) ---
    ax5 = fig.add_subplot(2, 3, (5, 6))
    domain_scores = defaultdict(lambda: Counter())
    for p in data:
        domain_scores[p['domain1']][p['score']] += 1
        domain_scores[p['domain2']][p['score']] += 1
    
    top_domains = sorted(dc.keys(), key=lambda x: -dc[x])[:6]
    x = np.arange(len(top_domains))
    width = 0.6
    bottom = np.zeros(len(top_domains))
    
    for s in [0, 1, 2, 3, 4]:
        vals = [domain_scores[d][s] for d in top_domains]
        ax5.bar(x, vals, width, bottom=bottom, label=f"Score {s}", color=SCORE_COLORS[s], edgecolor='white')
        bottom += np.array(vals)
    
    ax5.set_xticks(x)
    ax5.set_xticklabels(top_domains, fontsize=10)
    ax5.set_ylabel('Số lần xuất hiện')
    ax5.set_title('Phân bố Score theo Lĩnh vực (Top 6)', fontweight='bold')
    ax5.legend(fontsize=8, ncol=5, loc='upper right')
    ax5.spines['top'].set_visible(False)
    ax5.spines['right'].set_visible(False)
    
    # --- Thông tin tổng hợp ---
    total = len(data)
    fig.text(0.5, 0.01,
             f'Tổng: {total} cặp câu | {len(domains_sorted)} lĩnh vực | '
             f'Score TB: {sum(p["score"] for p in data)/total:.2f} | '
             f'Std: {np.std([p["score"] for p in data]):.2f}',
             ha='center', fontsize=11, fontstyle='italic',
             bbox=dict(boxstyle='round,pad=0.4', facecolor='lightyellow', alpha=0.9))
    
    plt.tight_layout(rect=[0, 0.03, 1, 0.95])
    plt.savefig(os.path.join(output_dir, '7_dashboard.png'))
    plt.close()
    print("  ✅ 7_dashboard.png")


def main():
    filepath = r'D:\MGEEMS\ViConSim_Dataset.xlsx'
    if not os.path.exists(filepath):
        print(f"Khong tim thay file: {filepath}")
        sys.exit(1)
    
    # Tạo thư mục output
    output_dir = r'D:\MGEEMS\dataset\Visualize'
    os.makedirs(output_dir, exist_ok=True)
    
    # Đọc data
    print("=" * 60)
    print("TRỰC QUAN HÓA DATASET ViConSim")
    print("=" * 60)
    print(f"\nĐọc file: {filepath}")
    data = read_dataset(filepath)
    print(f"Số cặp câu: {len(data)}")
    
    # Vẽ biểu đồ
    print(f"\nTạo biểu đồ (lưu vào {output_dir}/):")
    plot_score_distribution(data, output_dir)
    plot_domain_distribution(data, output_dir)
    plot_domain_pie(data, output_dir)
    plot_score_by_domain_type(data, output_dir)
    plot_heatmap_cross_domain(data, output_dir)
    plot_score_heatmap_by_domain(data, output_dir)
    plot_summary_dashboard(data, output_dir)
    
    print(f"\n{'=' * 60}")
    print(f"✅ HOÀN TẤT! {7} biểu đồ đã lưu trong: {output_dir}/")
    print(f"{'=' * 60}")


if __name__ == '__main__':
    main()