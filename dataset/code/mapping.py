

import csv
import sys
import os

try:
    from openpyxl import load_workbook, Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
except ImportError:
    print("Cần cài openpyxl: pip install openpyxl")
    sys.exit(1)


def build_gloss_map(concept_file, gloss_file):
    """Đọc 2 file line-by-line, tạo dict concept → gloss (unique)."""
    with open(concept_file, 'r', encoding='utf-8') as f:
        concepts = [line.strip() for line in f.readlines()]

    with open(gloss_file, 'r', encoding='utf-8') as f:
        glosses = [line.strip() for line in f.readlines()]

    print(f"File concepts: {len(concepts)} dòng")
    print(f"File glosses:  {len(glosses)} dòng")

    if len(concepts) != len(glosses):
        print(f"⚠️  Số dòng không khớp! Lấy min = {min(len(concepts), len(glosses))}")

    # Tạo mapping: concept → gloss (giữ bản đầu tiên)
    gloss_map = {}
    for i in range(min(len(concepts), len(glosses))):
        c = concepts[i].strip()
        g = glosses[i].strip()
        if c and g and c not in gloss_map:
            gloss_map[c] = g

    # Thêm mapping cho dạng readable (thay _ bằng space)
    extra = {}
    for c, g in gloss_map.items():
        readable = c.replace('_', ' ')
        if readable not in gloss_map:
            extra[readable] = g
        # Thêm dạng lowercase
        if c.lower() not in gloss_map:
            extra[c.lower()] = g
        if readable.lower() not in gloss_map:
            extra[readable.lower()] = g
    gloss_map.update(extra)

    print(f"Tổng mapping (bao gồm biến thể): {len(gloss_map)} entries")
    return gloss_map


def lookup_gloss(concept_text, gloss_map):
    """Tìm gloss cho một concept, thử nhiều biến thể."""
    if not concept_text:
        return ""

    text = concept_text.strip()

    # Thử nguyên bản
    if text in gloss_map:
        return gloss_map[text]

    # Thử thay space bằng _
    underscore = text.replace(' ', '_')
    if underscore in gloss_map:
        return gloss_map[underscore]

    # Thử lowercase
    if text.lower() in gloss_map:
        return gloss_map[text.lower()]

    if underscore.lower() in gloss_map:
        return gloss_map[underscore.lower()]

    # Thử bỏ dấu cách đầu/cuối, normalize
    for key in gloss_map:
        if key.lower().replace('_', ' ') == text.lower().replace('_', ' '):
            return gloss_map[key]

    return "[KHÔNG TÌM THẤY]"


def merge_glosses(concept_file, gloss_file, dataset_file):
    """Ghép gloss vào dataset ViConSim."""

    # 1. Tạo mapping
    print("=" * 60)
    print("BƯỚC 1: Tạo mapping concept → gloss")
    print("=" * 60)
    gloss_map = build_gloss_map(concept_file, gloss_file)

    # 2. Đọc dataset
    print("\n" + "=" * 60)
    print("BƯỚC 2: Đọc dataset ViConSim")
    print("=" * 60)
    wb = load_workbook(dataset_file)
    ws = wb.active
    print(f"Sheet: {ws.title}")
    print(f"Số dòng: {ws.max_row - 1} (không kể header)")
    print(f"Số cột hiện tại: {ws.max_column}")

    # Đọc header để xác định vị trí cột
    headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
    print(f"Headers: {headers}")

    # Tìm cột Concept 1 và Concept 2
    col_c1 = None
    col_c2 = None
    for i, h in enumerate(headers):
        if h and 'concept' in str(h).lower() and '1' in str(h):
            col_c1 = i + 1
        elif h and 'concept' in str(h).lower() and '2' in str(h):
            col_c2 = i + 1

    if not col_c1 or not col_c2:
        # Fallback: cột 2 = concept1, cột 3 = concept2
        col_c1 = 2
        col_c2 = 3
        print(f"⚠️  Không tìm thấy header chính xác, dùng mặc định: cột {col_c1} và {col_c2}")
    else:
        print(f"Cột Concept 1: {col_c1}, Cột Concept 2: {col_c2}")

    # 3. Thêm cột Gloss 1 và Gloss 2
    print("\n" + "=" * 60)
    print("BƯỚC 3: Ghép gloss vào dataset")
    print("=" * 60)

    # Xác định vị trí cột mới (sau cột cuối)
    col_g1 = ws.max_column + 1
    col_g2 = ws.max_column + 2

    # Style
    header_font = Font(bold=True, color="FFFFFF", size=11, name="Arial")
    header_fill = PatternFill("solid", fgColor="2F5496")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    wrap_align = Alignment(vertical="top", wrap_text=True)

    # Header cho cột mới
    for col, title in [(col_g1, "Gloss 1 (Sentence 1)"), (col_g2, "Gloss 2 (Sentence 2)")]:
        cell = ws.cell(row=1, column=col, value=title)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    # Ghép gloss cho từng dòng
    found = 0
    not_found = 0
    not_found_list = []

    for row in range(2, ws.max_row + 1):
        c1_val = ws.cell(row=row, column=col_c1).value
        c2_val = ws.cell(row=row, column=col_c2).value

        g1 = lookup_gloss(str(c1_val) if c1_val else "", gloss_map)
        g2 = lookup_gloss(str(c2_val) if c2_val else "", gloss_map)

        ws.cell(row=row, column=col_g1, value=g1).alignment = wrap_align
        ws.cell(row=row, column=col_g1).border = thin_border
        ws.cell(row=row, column=col_g2, value=g2).alignment = wrap_align
        ws.cell(row=row, column=col_g2).border = thin_border

        if g1 != "[KHÔNG TÌM THẤY]" and g2 != "[KHÔNG TÌM THẤY]":
            found += 1
        else:
            not_found += 1
            if g1 == "[KHÔNG TÌM THẤY]":
                not_found_list.append(c1_val)
            if g2 == "[KHÔNG TÌM THẤY]":
                not_found_list.append(c2_val)

    # Chỉnh độ rộng cột
    ws.column_dimensions[chr(64 + col_g1) if col_g1 <= 26 else 'H'].width = 60
    ws.column_dimensions[chr(64 + col_g2) if col_g2 <= 26 else 'I'].width = 60

    # 4. Lưu file
    output_file = dataset_file.replace('.xlsx', '_with_glosses.xlsx')
    if output_file == dataset_file:
        output_file = 'ViConSim_Dataset_with_glosses.xlsx'

    wb.save(output_file)

    # 5. Báo cáo
    print(f"\n✅ Ghép thành công: {found}/{found + not_found} cặp")
    if not_found > 0:
        print(f"⚠️  Không tìm thấy gloss: {not_found} cặp")
        unique_missing = list(set(not_found_list))[:10]
        print(f"   Một số concept thiếu: {unique_missing}")

    print(f"\n📁 File output: {output_file}")
    return output_file


if __name__ == '__main__':
    concept_file = r'D:\MGEEMS\dataset\concept.txt'
    gloss_file = r'D:\MGEEMS\dataset\gloss.txt'
    dataset_file = r'D:\MGEEMS\ViConSim_Dataset.xlsx'

    # Kiểm tra file tồn tại
    for f in [concept_file, gloss_file, dataset_file]:
        if not os.path.exists(f):
            print(f"Khong tim thay file: {f}")
            sys.exit(1)

    merge_glosses(concept_file, gloss_file, dataset_file)